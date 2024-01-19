#こちらが社内向けに開発したAWSの経費管理アプリになります。
#ローカル環境下にて動作するコードがこちら、後にも添付しておりますものが、AWS API Gatewayを通じて
#フロントサイドと通信し、AWS Lambda上で動作するように設計したコードになります。
################################################################################################
###################################    HOW TO USE    ###########################################
################################################################################################
################################################################################################

####################
import boto3
from boto3.dynamodb.conditions import Key
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from botocore.exceptions import ClientError, NoCredentialsError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

####################
#AWSアカウント名を入力。伏字にしております。以下のセッション項目も同様。
client_names = [...]  

# クライアントを生成するためのリスト内包表記を使用
sessions = [boto3.Session(profile_name=name) for name in client_names]
clients = [session.client('ce') for session in sessions]

# DynamoDBリソースの取得（最初のセッションを使用）
dynamodb = sessions[0].resource('dynamodb')

client_numbers = len(clients)#How many cliants are 
####################
# create a border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
####################
def date_info():
    # Get the current date
    now = datetime.now()
    # Get the date one year ago
    one_year_ago = now - timedelta(days=365)
    # Set the period for which to retrieve cost information
    start_date = one_year_ago.strftime('%Y-%m-%d')
    end_date = now.strftime('%Y-%m-%d')
    
    # Specify the account's start of use as 2022/11
    First_Month = datetime(2022,11,1)
    today = date.today()
    diff = relativedelta(today, First_Month)
    
    print(f'diff: {diff}')
    
    #You have to add one beecause the difference is less than quantity by one 
    month_qty = diff.years * 12 + diff.months + 1
    
    today = datetime.today()
    # Subtract the current day of the month to get last month
    last_month = today.replace(day=1) - timedelta(days=1)
    # Replace the day with 1 to get the first day of the month
    first_day_of_last_month = last_month.replace(day=1).strftime('%Y-%m-%d')
    
    return(now,
           one_year_ago,
           start_date,
           end_date,
           First_Month,
           today,
           month_qty,
           first_day_of_last_month
           )

def make_DDBtable(client_name):
    try:
        # Specify a non-existent table name
        table_description = dynamodb.meta.client.describe_table(TableName=client_name)
        table = dynamodb.Table(client_name)
    except ClientError as e:
        # Processing in case a ResourceNotFoundException occurs
        if e.response['Error']['Code'] == 'ResourceNotFoundException':
            table = dynamodb.create_table(
                TableName=client_name,
                KeySchema=[
                    {
                        'AttributeName': 'Date',
                        'KeyType': 'HASH'  # Partition key
                    },
                    {
                        'AttributeName': 'service_name',
                        'KeyType': 'RANGE'  # Sort key
                    }
                ],
                AttributeDefinitions=[
                    {
                        'AttributeName': 'Date',
                        'AttributeType': 'S'  # String type
                    },
                    {
                        'AttributeName': 'service_name',
                        'AttributeType': 'S'  # String type
                    },
                ],
                ProvisionedThroughput={
                    'ReadCapacityUnits': 5,
                    'WriteCapacityUnits': 5
                }
            )
            # Wait until the table is created
            table.meta.client.get_waiter('table_exists').wait(TableName=client_name)
    return(table)

def check_last_updated(client_name, ce_client, table,now, one_year_ago, start_date, end_date, first_day_of_last_month):
    
    # Initialize a dictionary to store the total cost for each month
    monthly_total = {}
    
    # Get the latest cost information from the DynamoDB table
    response = table.query(
        KeyConditionExpression=boto3.dynamodb.conditions.Key('Date').eq(first_day_of_last_month),
        Limit=1,
    )
    
    # If the DynamoDB table does not have the latest cost information, call the API to get the cost information
    if not response['Items']:
        try:
            print("Warning")
            service_billing = ce_client.get_cost_and_usage(
                TimePeriod={
                    'Start': start_date,
                    'End': end_date
                },
                Granularity='MONTHLY',
                Metrics=[
                    'UnblendedCost'
                ],
                GroupBy=[
                    {
                        'Type': 'DIMENSION',
                        'Key': 'SERVICE'
                    }
                ]
            )
            
            
            # Iterate over the results returned from the API
            for result_by_time in service_billing['ResultsByTime']:
                for group in result_by_time['Groups']:
                    
                    #put the responce to the table
                    service_billing={
                        'Date': result_by_time['TimePeriod']['Start'],
                        'service_name': group['Keys'][0],
                        'billing': group['Metrics']['UnblendedCost']['Amount'],
                    }
                    table.put_item(Item=service_billing)
                    
                    service_cost = float(group['Metrics']['UnblendedCost']['Amount'])
                    
                    month = result_by_time['TimePeriod']['Start']
                    
                    #To calucuate monthly subtotal
                    # If the month already exists in the dictionary, add the cost
                    if month in monthly_total:
                        monthly_total[month] += service_cost
                    # If the month does not exist in the dictionary, create a new entry
                    else:
                        monthly_total[month] = service_cost
                        
            # After looping, the monthly costs are in monthly_costs{}, so store these in the Table
            for Date, billing in monthly_total.items():
                    monthly_total={
                                    'Date': Date,
                                    'service_name': 'Total',
                                    'billing': str(billing)
                    }
                    table.put_item(Item=monthly_total)
                    
            return False
                        
        except NoCredentialsError:
            print("No AWS credentials were found.")

def get_data_from_DDB(table, First_Month, today, month_qty):
    #Initialize First_Month at each loop, because "for client_name", "ce_client in zip(client_names, clients):" exsisting
    First_Month = datetime(2022, 11, 1)
    each_cost_info = []  # List to store all the results
    
    for i in range(month_qty):
        response = table.query(
            KeyConditionExpression=boto3.dynamodb.conditions.Key('Date').eq(First_Month.strftime('%Y-%m-%d')),
            ScanIndexForward=True
        )
        each_cost_info.extend(response['Items'])
        First_Month += relativedelta(months=1)
        
    return each_cost_info
    
def Write_to_Excel(all_cost_info):
    
    write_Account_Summary(all_cost_info)
    write_Service_Detail(all_cost_info)
    
def write_Account_Summary(all_cost_info):
    total_billing = []
    Account_Summary = {}
    
    for key, value in all_cost_info.items():
        total_billing = [item for item in value if item.get('service_name') == 'Total']#Total だけを抽出
        Account_Summary[key] = total_billing #再度入れる
                
    df = pd.concat({k: pd.DataFrame(v) for k, v in Account_Summary.items()}, names=['Development Phase & Clouds', 'Date_billing'])
    df.reset_index(level='Development Phase & Clouds', inplace=True)
    # Convert billing from string to numeric
    df['billing'] = pd.to_numeric(df['billing'])
    pivot_table = df.pivot(index='Development Phase & Clouds', columns='Date', values='billing')
    
    pivot_table.fillna(0, inplace=True)
    pivot_table.loc['Total',:] = pivot_table.sum(axis=0, numeric_only=True)
    # Extract the "Total" row to calculate the cumulative sum
    total_row_cumulative = pivot_table.loc['Total'].cumsum()
    pivot_table.loc[:,'Total'] = pivot_table.sum(axis=1, numeric_only=True)
    #round
    pivot_table = pivot_table.round(2)
    
    with pd.ExcelWriter('C:\\AWS\\my_dataframe.xlsx', engine='openpyxl') as writer:
        pivot_table.to_excel(writer, sheet_name='Development Phase & Clouds', index=True)

    wb = load_workbook('C:\\AWS\\my_dataframe.xlsx')
    ws = wb['Development Phase & Clouds']
    
    # Extract the "Total" row to calculate the cumulative sum
    # total_row_cumulative = pivot_table.loc['Total'].cumsum()
    
    # グラフを作成
    plt.figure(figsize= (12, 6))  # グラフのサイズを指定
    total_row_cumulative.plot(kind='line')

    
    # グラフのタイトルとラベルを設定
    plt.title('Total Billing Over Time')
    plt.xlabel('Date')
    plt.ylabel('Total Billing(USD)')
    
    # Y軸の目盛りを設定
    min_val = total_row_cumulative.min()  # データの最小値
    max_val = total_row_cumulative.max()  # データの最大値
    yticks = np.linspace(min_val, max_val, num=10)  # 最小値から最大値まで均等に10個の目盛りを生成
    plt.yticks(yticks)
    plt.grid(True)
    # グラフを一時的な画像ファイルとして保存
    plt.savefig('temp_plot.png')
    # 画像ファイルをロードしてワークシートに追加
    img = Image('temp_plot.png')
    ws.add_image(img, 'A26')  # 位置は適宜調整
        
    # Automatically adjust the cell width and height
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    # apply border to cells
    for row in ws:
        for cell in row:
            cell.border = thin_border
    
    # Save the Excel file
    wb.save("C:\\AWS\\my_dataframe.xlsx")
    
def write_Service_Detail(all_cost_info):
    Service_Detail = []
    Grouped_Detail = {}
    
    for key, value in all_cost_info.items():
        Service_Detail = [item for item in value if item.get('service_name') != 'Total']#Total以外を抽出
        Grouped_Detail[key] = Service_Detail # Insert again
        
    for client_name_k, Service_Date_billing_v in Grouped_Detail.items():
        
        df = pd.concat({client_name_k: pd.DataFrame(Service_Date_billing_v)}, names=['Development Phase & Clouds', 'Service_Date_billing'])
        df.reset_index(level='Development Phase & Clouds', inplace=True)
        # Convert billing from string to numeric
        df['billing'] = pd.to_numeric(df['billing'])
        pivot_table = df.pivot(index='service_name', columns='Date', values='billing')
        
        pivot_table.fillna(0, inplace=True)
        pivot_table.loc['Total',:] = pivot_table.sum(axis=0, numeric_only=True)
        pivot_table.loc[:,'Total'] = pivot_table.sum(axis=1, numeric_only=True)
        #round
        pivot_table = pivot_table.round(2)
        # Load workbook to prevent overwriting
        book = load_workbook('C:\\AWS\\my_dataframe.xlsx')
    
        with pd.ExcelWriter('C:\\AWS\\my_dataframe.xlsx', engine='openpyxl') as writer:
            # Open the workbook to prevent overwriting
            writer.book = book
            pivot_table.to_excel(writer, sheet_name= client_name_k, index=True)

        wb = load_workbook('C:\\AWS\\my_dataframe.xlsx')
        ws = wb[client_name_k]

        # Automatically adjust the cell width and height
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        # apply border to cells
        for row in ws:
            for cell in row:
                cell.border = thin_border
        # Save the Excel file
        wb.save("C:\\AWS\\my_dataframe.xlsx")
    
    wb.close()
    
def main():
    
    (now,
    one_year_ago,
    start_date,
    end_date,
    First_Month,
    today,
    month_qty,
    first_day_of_last_month
    ) = date_info()
    
    # Dictionary to store all cost information
    all_cost_info = {}

    #To check updated status at each acount
    for client_name, ce_client in zip(client_names, clients):
        
        table = make_DDBtable(client_name)
        
        check_last_updated(client_name, ce_client, table, now, one_year_ago, start_date, end_date, first_day_of_last_month)
        
        each_cost_info = get_data_from_DDB(table, First_Month, today, month_qty)
    
        all_cost_info[client_name] = each_cost_info
        
    # print(all_cost_info)
    Write_to_Excel(all_cost_info)
    

if __name__ == "__main__":
    main()