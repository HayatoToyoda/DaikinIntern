#AWS API Gatewayを通じて、フロントサイドと通信し、AWS Lambda上で動作するように設計したコードになります。
################################################################################################
###################################    HOW TO USE    ###########################################
################################################################################################
################################################################################################

####################
####################
from io import BytesIO
import base64
import boto3
from boto3.dynamodb.conditions import Key
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from botocore.exceptions import ClientError, NoCredentialsError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np

####################
client_names = [
                'a', 
                'b', 
                'c', 
                'd', 
                'e', 
                'f', 
                'g', 
                'h', 
                'i', 
                'j'
                ]

gpf_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client0= gpf_beta.client('ce')

preprod_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client1= preprod_beta.client('ce')

serviceintegration_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key="",
    region_name='us-west-2'#specify the region because this account used for DDB
)
client2= serviceintegration_beta.client('ce')

hero_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client3= hero_beta.client('ce')


commisioning_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client4= commisioning_beta.client('ce')


Leak_Detection_beta = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client5= Leak_Detection_beta.client('ce')


hero_prod = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client6= hero_prod.client('ce')


gpf_prod = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client7= gpf_prod.client('ce')


commisioning_prod = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client8= commisioning_prod.client('ce')


Leak_Detection_prod = boto3.Session(
    aws_access_key_id="",
    aws_secret_access_key=""
)
client9= Leak_Detection_prod.client('ce')


dynamodb = serviceintegration_beta.resource('dynamodb')



#table = dynamodb.Table('cost_explorer-table')

clients = [client0, client1, client2, client3, client4, client5, client6, client7, client8, client9]

client_numbers = len(clients)#How many cliants are 
####################
# create a border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
####################
def date_info():
    # Get the current date
    now = datetime.now()
    # Get the date one year ago
    one_year_ago = now - timedelta(days=365)
    # Set the period for which to retrieve cost information
    start_date = one_year_ago.strftime('%Y-%m-%d')
    end_date = now.strftime('%Y-%m-%d')
    
    # Specify the account's start of use as 2022/11
    First_Month = datetime(2022,11,1)
    today = date.today()
    diff = relativedelta(today, First_Month)
    
    print(f'diff: {diff}')
    
    month_diff = diff.years * 12 + diff.months
    
    today = datetime.today()
    # Subtract the current day of the month to get last month
    last_month = today.replace(day=1) - timedelta(days=1)
    # Replace the day with 1 to get the first day of the month
    first_day_of_last_month = last_month.replace(day=1).strftime('%Y-%m-%d')
    
    return(now,
           one_year_ago,
           start_date,
           end_date,
           First_Month,
           today,
           month_diff,
           first_day_of_last_month
           )

def make_DDBtable(client_name):
    try:
        # Specify a non-existent table name
        table_description = dynamodb.meta.client.describe_table(TableName=client_name)
        table = dynamodb.Table(client_name)
    except ClientError as e:
        # Processing in case a ResourceNotFoundException occurs
        if e.response['Error']['Code'] == 'ResourceNotFoundException':
            table = dynamodb.create_table(
                TableName=client_name,
                KeySchema=[
                    {
                        'AttributeName': 'Date',
                        'KeyType': 'HASH'  # Partition key
                    },
                    {
                        'AttributeName': 'service_name',
                        'KeyType': 'RANGE'  # Sort key
                    }
                ],
                AttributeDefinitions=[
                    {
                        'AttributeName': 'Date',
                        'AttributeType': 'S'  # String type
                    },
                    {
                        'AttributeName': 'service_name',
                        'AttributeType': 'S'  # String type
                    },
                ],
                ProvisionedThroughput={
                    'ReadCapacityUnits': 5,
                    'WriteCapacityUnits': 5
                }
            )
            # Wait until the table is created
            table.meta.client.get_waiter('table_exists').wait(TableName=client_name)
    return(table)

def check_last_updated(client_name, ce_client, table,now, one_year_ago, start_date, end_date, first_day_of_last_month):
    
    # Initialize a dictionary to store the total cost for each month
    monthly_total = {}
    
    # Get the latest cost information from the DynamoDB table
    response = table.query(
        KeyConditionExpression=boto3.dynamodb.conditions.Key('Date').eq(first_day_of_last_month),
        Limit=1,
    )
    
    # If the DynamoDB table does not have the latest cost information, call the API to get the cost information
    if not response['Items']:
        try:
            print("Warning")
            service_billing = ce_client.get_cost_and_usage(
                TimePeriod={
                    'Start': start_date,
                    'End': end_date
                },
                Granularity='MONTHLY',
                Metrics=[
                    'UnblendedCost'
                ],
                GroupBy=[
                    {
                        'Type': 'DIMENSION',
                        'Key': 'SERVICE'
                    }
                ]
            )
            
            
            # Iterate over the results returned from the API
            for result_by_time in service_billing['ResultsByTime']:
                for group in result_by_time['Groups']:
                    
                    #put the responce to the table
                    service_billing={
                        'Date': result_by_time['TimePeriod']['Start'],
                        'service_name': group['Keys'][0],
                        'billing': group['Metrics']['UnblendedCost']['Amount'],
                    }
                    table.put_item(Item=service_billing)
                    
                    service_cost = float(group['Metrics']['UnblendedCost']['Amount'])
                    
                    month = result_by_time['TimePeriod']['Start']
                    
                    #To calucuate monthly subtotal
                    # If the month already exists in the dictionary, add the cost
                    if month in monthly_total:
                        monthly_total[month] += service_cost
                    # If the month does not exist in the dictionary, create a new entry
                    else:
                        monthly_total[month] = service_cost
                        
            # After looping, the monthly costs are in monthly_costs{}, so store these in the Table
            for Date, billing in monthly_total.items():
                    monthly_total={
                                    'Date': Date,
                                    'service_name': 'Total',
                                    'billing': str(billing)
                    }
                    table.put_item(Item=monthly_total)
                    
            return False
                        
        except NoCredentialsError:
            print("No AWS credentials were found.")

def get_data_from_DDB(table, First_Month, today, month_diff):
    #Initialize First_Month at each loop, because "for client_name", "ce_client in zip(client_names, clients):" exsisting
    First_Month = datetime(2022, 11, 1)
    each_cost_info = []  # List to store all the results
    
    for i in range(month_diff):
        response = table.query(
            KeyConditionExpression=boto3.dynamodb.conditions.Key('Date').eq(First_Month.strftime('%Y-%m-%d')),
            ScanIndexForward=True
        )
        each_cost_info.extend(response['Items'])
        First_Month += relativedelta(months=1)
        
    return each_cost_info
    
def Write_to_Excel(all_cost_info):
    
    wb_account = write_Account_Summary(all_cost_info)
    wb_service = write_Service_Detail(all_cost_info, wb_account)
    return wb_service
    
def write_Account_Summary(all_cost_info):
    total_billing = []
    Account_Summary = {}
    
    for key, value in all_cost_info.items():
        total_billing = [item for item in value if item.get('service_name') == 'Total']#Total だけを抽出
        Account_Summary[key] = total_billing #再度入れる
                
    df = pd.concat({k: pd.DataFrame(v) for k, v in Account_Summary.items()}, names=['Development Phase & Clouds', 'Date_billing'])
    df.reset_index(level='Development Phase & Clouds', inplace=True)
    # Convert billing from string to numeric
    df['billing'] = pd.to_numeric(df['billing'])
    pivot_table = df.pivot(index='Development Phase & Clouds', columns='Date', values='billing')
    
    pivot_table.fillna(0, inplace=True)
    pivot_table.loc['Total',:] = pivot_table.sum(axis=0, numeric_only=True)
    # Extract the "Total" row to calculate the cumulative sum
    total_row_cumulative = pivot_table.loc['Total'].cumsum()
    pivot_table.loc[:,'Total'] = pivot_table.sum(axis=1, numeric_only=True)
    #round
    pivot_table = pivot_table.round(2)
    
    # Create new and write
    with pd.ExcelWriter('/tmp/my_dataframe.xlsx', engine='openpyxl') as writer:
        pivot_table.to_excel(writer, sheet_name='Development Phase & Clouds', index=True)
        # Make the sheet visible
        writer.sheets['Development Phase & Clouds'].sheet_state = 'visible'

    wb = load_workbook('/tmp/my_dataframe.xlsx')
    ws = wb['Development Phase & Clouds']

     # Automatically adjust the cell width and height
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    # apply border to cells
    for row in ws:
        for cell in row:
           cell.border = thin_border
    
    # Save the Excel file
    wb.save('/tmp/my_dataframe.xlsx')
    return wb
    
def write_Service_Detail(all_cost_info, wb):
    Service_Detail = []
    Grouped_Detail = {}
    
    for key, value in all_cost_info.items():
        Service_Detail = [item for item in value if item.get('service_name') != 'Total']#Total以外を抽出
        Grouped_Detail[key] = Service_Detail # Insert again
        
    for client_name_k, Service_Date_billing_v in Grouped_Detail.items():
        
        df = pd.concat({client_name_k: pd.DataFrame(Service_Date_billing_v)}, names=['Development Phase & Clouds', 'Service_Date_billing'])
        df.reset_index(level='Development Phase & Clouds', inplace=True)
        #billing が文字列なのでnumericへ
        df['billing'] = pd.to_numeric(df['billing'])
        pivot_table = df.pivot(index='service_name', columns='Date', values='billing')
        
        pivot_table.fillna(0, inplace=True)
        pivot_table.loc['Total',:] = pivot_table.sum(axis=0, numeric_only=True)
        pivot_table.loc[:,'Total'] = pivot_table.sum(axis=1, numeric_only=True)
        #round
        pivot_table = pivot_table.round(2)
        # Load workbook to prevent overwriting
        book = load_workbook('/tmp/my_dataframe.xlsx')
    
        with pd.ExcelWriter('/tmp/my_dataframe.xlsx', engine='openpyxl') as writer:
            # Open the workbook to prevent overwriting
            writer.book = book
            pivot_table.to_excel(writer, sheet_name= client_name_k, index=True)
            writer.sheets[client_name_k].sheet_state = 'visible'

        wb = load_workbook('/tmp/my_dataframe.xlsx')
        ws = wb[client_name_k]

        # Automatically adjust the cell width and height
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        # apply border to cells
        for row in ws:
            for cell in row:
                cell.border = thin_border
        # Save the Excel file
        wb.save('/tmp/my_dataframe.xlsx')
    
    #wb.close()
    return wb
    
def main():
    
    (now,
    one_year_ago,
    start_date,
    end_date,
    First_Month,
    today,
    month_diff,
    first_day_of_last_month
    ) = date_info()
    
    # Dictionary to store all cost information
    all_cost_info = {}

    should_break = False

    #To check updated status at each acount
    for client_name, ce_client in zip(client_names, clients):

        if should_break:
            break

        table = make_DDBtable(client_name)
        
        check_last_updated(client_name, ce_client, table, now, one_year_ago, start_date, end_date, first_day_of_last_month)
        
        each_cost_info = get_data_from_DDB(table, First_Month, today, month_diff)
    
        all_cost_info[client_name] = each_cost_info

        should_break = True

    wb = Write_to_Excel(all_cost_info)

    binary_data_of_Excel = BytesIO()
    wb.save(binary_data_of_Excel)
    
    print(binary_data_of_Excel)
    
    # return {
    #     'statusCode': 200,
    #     'headers': {
    #         'Content-Type': 'application/octet-stream',
    #         'Content-Disposition': f'attachment; filename={FILE_NAME}'
    #     },
    #     'body': base64.b64encode(binary_data_of_Excel).decode('utf-8'),
    #     'isBase64Encoded': True
    # }
    



if __name__ == "__main__":
    main()